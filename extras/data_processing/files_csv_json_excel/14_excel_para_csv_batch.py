"""
Conversao em lote de varias planilhas Excel para CSV

O que este script demonstra: percorrer multiplos arquivos .xlsx em uma pasta e
converter cada aba de cada arquivo em um CSV separado.
Quando usar: quando um processo downstream (ex: import para banco, ferramenta de BI)
so aceita CSV, mas a entrada e fornecida como varias planilhas Excel.
"""

import os
import tempfile

from openpyxl import Workbook, load_workbook
import csv


def gerar_planilhas_de_exemplo(pasta: str) -> list[str]:
    """Gera 2 arquivos .xlsx, um deles com 2 abas, para simular um lote real."""
    caminhos = []

    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = "Dados"
    ws1.append(["produto", "preco"])
    ws1.append(["Mouse", 49.9])
    caminho1 = os.path.join(pasta, "relatorio1.xlsx")
    wb1.save(caminho1)
    caminhos.append(caminho1)

    wb2 = Workbook()
    ws2a = wb2.active
    ws2a.title = "Janeiro"
    ws2a.append(["dia", "vendas"])
    ws2a.append([1, 100])
    ws2b = wb2.create_sheet("Fevereiro")
    ws2b.append(["dia", "vendas"])
    ws2b.append([1, 120])
    caminho2 = os.path.join(pasta, "relatorio2.xlsx")
    wb2.save(caminho2)
    caminhos.append(caminho2)

    return caminhos


def converter_excel_para_csv_batch(caminhos_xlsx: list[str], pasta_saida: str) -> list[str]:
    """
    Para cada arquivo Excel, gera um CSV por aba. Nome do CSV de saida combina o
    nome do arquivo original com o nome da aba, evitando colisao quando duas
    planilhas diferentes tem abas com o mesmo nome.
    """
    arquivos_gerados = []
    for caminho_xlsx in caminhos_xlsx:
        nome_base = os.path.splitext(os.path.basename(caminho_xlsx))[0]
        wb = load_workbook(caminho_xlsx, read_only=True, data_only=True)
        for nome_aba in wb.sheetnames:
            ws = wb[nome_aba]
            caminho_csv = os.path.join(pasta_saida, f"{nome_base}__{nome_aba}.csv")
            with open(caminho_csv, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                for linha in ws.iter_rows(values_only=True):
                    writer.writerow(linha)
            arquivos_gerados.append(caminho_csv)
        wb.close()
    return arquivos_gerados


if __name__ == "__main__":
    tmp_dir = tempfile.mkdtemp()
    pasta_entrada = os.path.join(tmp_dir, "entrada")
    pasta_saida = os.path.join(tmp_dir, "saida")
    os.makedirs(pasta_entrada)
    os.makedirs(pasta_saida)

    try:
        caminhos_xlsx = gerar_planilhas_de_exemplo(pasta_entrada)
        csvs_gerados = converter_excel_para_csv_batch(caminhos_xlsx, pasta_saida)

        print("CSVs gerados:")
        for caminho in csvs_gerados:
            print(" -", os.path.basename(caminho))

        # relatorio1 tem 1 aba, relatorio2 tem 2 abas = 3 CSVs no total
        assert len(csvs_gerados) == 3
        nomes_gerados = {os.path.basename(c) for c in csvs_gerados}
        assert "relatorio2__Fevereiro.csv" in nomes_gerados
        print("OK: conversao em lote de Excel para CSV concluida.")
    finally:
        for pasta in (pasta_entrada, pasta_saida):
            for nome_arquivo in os.listdir(pasta):
                os.remove(os.path.join(pasta, nome_arquivo))
            os.rmdir(pasta)
        os.rmdir(tmp_dir)
