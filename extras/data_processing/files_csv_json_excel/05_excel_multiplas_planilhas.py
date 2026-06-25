"""
Leitura e escrita de multiplas planilhas (abas) com openpyxl

O que este script demonstra: criar um arquivo Excel com varias abas (worksheets) e
depois ler cada aba de volta, navegando pelo workbook.
Quando usar: quando dados logicamente relacionados mas distintos (ex: vendas por
regiao, ou tabelas de dimensao/fato) precisam viver em um unico arquivo .xlsx.
"""

import os
import tempfile

from openpyxl import Workbook, load_workbook


def criar_workbook_multi_abas(caminho: str) -> None:
    """Cria um workbook com 3 abas: Vendas, Clientes e Resumo."""
    wb = Workbook()

    # a aba padrao criada automaticamente e renomeada em vez de descartada
    ws_vendas = wb.active
    ws_vendas.title = "Vendas"
    ws_vendas.append(["id_venda", "produto", "valor"])
    ws_vendas.append([1, "Mouse", 50.0])
    ws_vendas.append([2, "Teclado", 120.0])

    ws_clientes = wb.create_sheet("Clientes")
    ws_clientes.append(["id_cliente", "nome"])
    ws_clientes.append([1, "Ana"])
    ws_clientes.append([2, "Bruno"])

    ws_resumo = wb.create_sheet("Resumo")
    ws_resumo.append(["metrica", "valor"])
    ws_resumo.append(["total_vendas", 170.0])

    wb.save(caminho)


def ler_todas_as_abas(caminho: str) -> dict[str, list[tuple]]:
    """Le o workbook e retorna um dict {nome_da_aba: linhas}."""
    wb = load_workbook(caminho, read_only=True, data_only=True)
    dados = {}
    for nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
        dados[nome_aba] = [linha for linha in ws.iter_rows(values_only=True)]
    wb.close()
    return dados


if __name__ == "__main__":
    tmp_dir = tempfile.mkdtemp()
    caminho_xlsx = os.path.join(tmp_dir, "multi_abas.xlsx")

    try:
        criar_workbook_multi_abas(caminho_xlsx)

        dados = ler_todas_as_abas(caminho_xlsx)
        for aba, linhas in dados.items():
            print(f"--- {aba} ---")
            for linha in linhas:
                print(linha)

        assert set(dados.keys()) == {"Vendas", "Clientes", "Resumo"}
        assert dados["Vendas"][0] == ("id_venda", "produto", "valor")
        assert dados["Vendas"][1][1] == "Mouse"
        print("OK: multiplas abas criadas e lidas corretamente.")
    finally:
        if os.path.exists(caminho_xlsx):
            os.remove(caminho_xlsx)
        os.rmdir(tmp_dir)
