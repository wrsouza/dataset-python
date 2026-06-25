"""
Formatacao condicional e estilo de celulas com openpyxl

O que este script demonstra: aplicar fontes, cores de fundo e formatacao condicional
(via PatternFill conforme regra) em celulas de uma planilha Excel.
Quando usar: ao gerar relatorios Excel que precisam destacar visualmente valores
(ex: vermelho para negativo, verde para acima da meta) sem o usuario abrir filtros.
"""

import os
import tempfile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.formatting.rule import CellIsRule


VERMELHO = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
VERDE = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")


def criar_planilha_formatada(caminho: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"

    cabecalho = ["produto", "meta", "realizado", "diferenca"]
    ws.append(cabecalho)

    # cabecalho em negrito, fundo cinza e centralizado
    for col_idx, _ in enumerate(cabecalho, start=1):
        celula = ws.cell(row=1, column=col_idx)
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        celula.alignment = Alignment(horizontal="center")

    linhas = [
        ("Mouse", 100, 120, 20),
        ("Teclado", 80, 60, -20),
        ("Monitor", 50, 50, 0),
    ]
    for linha in linhas:
        ws.append(linha)

    # formatacao condicional: coluna D (diferenca) negativa em vermelho, positiva em verde.
    # CellIsRule compara contra um literal fixo ("0"); range comeca na linha 2 para nao
    # atingir o cabecalho.
    intervalo = "D2:D4"
    ws.conditional_formatting.add(
        intervalo, CellIsRule(operator="lessThan", formula=["0"], fill=VERMELHO)
    )
    ws.conditional_formatting.add(
        intervalo, CellIsRule(operator="greaterThan", formula=["0"], fill=VERDE)
    )

    # largura de coluna automatica simplificada (openpyxl nao calcula isso nativamente)
    for col_idx, titulo in enumerate(cabecalho, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(
            12, len(titulo) + 4
        )

    wb.save(caminho)


def verificar_formatacao(caminho: str) -> None:
    wb = load_workbook(caminho)
    ws = wb["Resultados"]

    cabecalho_celula = ws["A1"]
    assert cabecalho_celula.font.bold, "Cabecalho deveria estar em negrito"

    regras = ws.conditional_formatting
    # apenas garante que ao menos uma regra foi registrada no range esperado
    ranges_registrados = [str(rng) for rng in regras]
    assert any("D2:D4" in r for r in ranges_registrados), "Regra condicional ausente"
    wb.close()


if __name__ == "__main__":
    tmp_dir = tempfile.mkdtemp()
    caminho_xlsx = os.path.join(tmp_dir, "formatado.xlsx")

    try:
        criar_planilha_formatada(caminho_xlsx)
        verificar_formatacao(caminho_xlsx)
        print("OK: planilha com estilos e formatacao condicional criada e validada.")
    finally:
        if os.path.exists(caminho_xlsx):
            os.remove(caminho_xlsx)
        os.rmdir(tmp_dir)
