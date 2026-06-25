"""
Gerar relatorio Excel formatado a partir de dados CSV

O que este script demonstra: ler dados tabulares de um CSV e gerar um arquivo Excel
com cabecalho estilizado, larguras de coluna ajustadas, formato numerico de moeda e
uma linha de totais - um relatorio pronto para ser aberto por um usuario de negocio.
Quando usar: ao final de um pipeline de dados, quando o publico final consome
Excel (nao CSV cru) e espera um documento legivel e ja formatado.
"""

import csv
import io
import os
import tempfile

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def csv_para_excel_formatado(texto_csv: str, caminho_saida: str, coluna_valor: str) -> None:
    leitor = csv.DictReader(io.StringIO(texto_csv))
    linhas = list(leitor)
    colunas = leitor.fieldnames or []

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio"

    # cabecalho
    for col_idx, nome_coluna in enumerate(colunas, start=1):
        celula = ws.cell(row=1, column=col_idx, value=nome_coluna)
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
        celula.alignment = Alignment(horizontal="center")

    # dados - converte a coluna de valor para float para permitir formato de moeda e soma
    total = 0.0
    for row_idx, linha in enumerate(linhas, start=2):
        for col_idx, nome_coluna in enumerate(colunas, start=1):
            valor_bruto = linha[nome_coluna]
            if nome_coluna == coluna_valor:
                valor = float(valor_bruto)
                total += valor
                celula = ws.cell(row=row_idx, column=col_idx, value=valor)
                celula.number_format = '#,##0.00'
            else:
                ws.cell(row=row_idx, column=col_idx, value=valor_bruto)

    # linha de total ao final
    linha_total = len(linhas) + 2
    indice_coluna_valor = colunas.index(coluna_valor) + 1
    ws.cell(row=linha_total, column=1, value="TOTAL").font = Font(bold=True)
    celula_total = ws.cell(row=linha_total, column=indice_coluna_valor, value=total)
    celula_total.font = Font(bold=True)
    celula_total.number_format = '#,##0.00'

    # largura de coluna baseada no maior conteudo da coluna (aproximacao simples)
    for col_idx, nome_coluna in enumerate(colunas, start=1):
        maior_tamanho = max(
            [len(str(nome_coluna))]
            + [len(str(linha[nome_coluna])) for linha in linhas]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = maior_tamanho + 4

    ws.freeze_panes = "A2"  # mantem o cabecalho visivel ao rolar

    wb.save(caminho_saida)


if __name__ == "__main__":
    csv_exemplo = (
        "produto,quantidade,valor_total\n"
        "Mouse,10,499.00\n"
        "Teclado,5,600.00\n"
        "Monitor,2,1600.00\n"
    )

    tmp_dir = tempfile.mkdtemp()
    caminho_xlsx = os.path.join(tmp_dir, "relatorio_formatado.xlsx")

    try:
        csv_para_excel_formatado(csv_exemplo, caminho_xlsx, coluna_valor="valor_total")

        # validacao basica: reabre o arquivo e confere estrutura e total calculado
        from openpyxl import load_workbook

        wb = load_workbook(caminho_xlsx)
        ws = wb["Relatorio"]

        assert ws["A1"].value == "produto"
        assert ws["A1"].font.bold is True
        valor_total_calculado = ws.cell(row=5, column=3).value
        assert abs(valor_total_calculado - 2699.0) < 1e-6, "Total deveria ser 499+600+1600"
        wb.close()

        print(f"Relatorio gerado em: {caminho_xlsx}")
        print("OK: relatorio Excel formatado gerado e validado com sucesso.")
    finally:
        if os.path.exists(caminho_xlsx):
            os.remove(caminho_xlsx)
        os.rmdir(tmp_dir)
