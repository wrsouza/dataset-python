"""Concrete Report Builders: PDF, Excel, CSV."""
from __future__ import annotations

import csv
import io
from typing import Self

from report_builder.domain.entities import (
    ChartSpec,
    DataTable,
    Report,
    ReportFormat,
)
from report_builder.domain.interfaces import ReportBuilder


class CSVReportBuilder(ReportBuilder):
    """Builds a CSV report using the standard library — no extra deps.

    SRP: only responsible for CSV serialisation.
    """

    def __init__(self) -> None:
        self._title: str = ""
        self._headers: list[str] = []
        self._tables: list[DataTable] = []
        self._footer: str = ""

    def set_title(self, title: str) -> Self:
        self._title = title
        return self

    def add_header(self, text: str) -> Self:
        self._headers.append(text)
        return self

    def add_data_table(self, table: DataTable) -> Self:
        self._tables.append(table)
        return self

    def add_chart(self, chart: ChartSpec) -> Self:
        # CSV cannot render charts; silently accepted to honour the interface.
        return self

    def add_footer(self, text: str) -> Self:
        self._footer = text
        return self

    def build(self) -> Report:
        buf = io.StringIO()
        writer = csv.writer(buf)

        if self._title:
            writer.writerow([self._title])

        for header in self._headers:
            writer.writerow([header])

        for table in self._tables:
            if table.title:
                writer.writerow([table.title])
            writer.writerow(table.headers)
            for row in table.rows:
                writer.writerow(row)
            writer.writerow([])

        if self._footer:
            writer.writerow([self._footer])

        return Report(
            title=self._title,
            headers=self._headers,
            data_tables=self._tables,
            footer=self._footer,
            format=ReportFormat.CSV,
            output=buf.getvalue().encode("utf-8"),
            content_type="text/csv",
            filename=f"{self._title or 'report'}.csv",
        )


class ExcelReportBuilder(ReportBuilder):
    """Builds an Excel (.xlsx) report using openpyxl.

    SRP: only responsible for Excel serialisation.
    """

    def __init__(self) -> None:
        self._title: str = ""
        self._headers: list[str] = []
        self._tables: list[DataTable] = []
        self._charts: list[ChartSpec] = []
        self._footer: str = ""

    def set_title(self, title: str) -> Self:
        self._title = title
        return self

    def add_header(self, text: str) -> Self:
        self._headers.append(text)
        return self

    def add_data_table(self, table: DataTable) -> Self:
        self._tables.append(table)
        return self

    def add_chart(self, chart: ChartSpec) -> Self:
        self._charts.append(chart)
        return self

    def add_footer(self, text: str) -> Self:
        self._footer = text
        return self

    def build(self) -> Report:
        import openpyxl
        from openpyxl.styles import Font

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self._title or "Report"  # type: ignore[union-attr]

        current_row = 1

        if self._title:
            ws.cell(row=current_row, column=1, value=self._title).font = Font(bold=True, size=14)  # type: ignore[union-attr]
            current_row += 2

        for header in self._headers:
            ws.cell(row=current_row, column=1, value=header)  # type: ignore[union-attr]
            current_row += 1

        for table in self._tables:
            if table.title:
                ws.cell(row=current_row, column=1, value=table.title).font = Font(bold=True)  # type: ignore[union-attr]
                current_row += 1

            for col_idx, col_header in enumerate(table.headers, start=1):
                ws.cell(row=current_row, column=col_idx, value=col_header).font = Font(bold=True)  # type: ignore[union-attr]
            current_row += 1

            for row in table.rows:
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row=current_row, column=col_idx, value=value)  # type: ignore[union-attr]
                current_row += 1

            current_row += 1

        if self._footer:
            ws.cell(row=current_row, column=1, value=self._footer)  # type: ignore[union-attr]

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return Report(
            title=self._title,
            headers=self._headers,
            data_tables=self._tables,
            charts=self._charts,
            footer=self._footer,
            format=ReportFormat.EXCEL,
            output=buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"{self._title or 'report'}.xlsx",
        )


class PDFReportBuilder(ReportBuilder):
    """Builds a PDF report using reportlab.

    SRP: only responsible for PDF serialisation.
    """

    def __init__(self) -> None:
        self._title: str = ""
        self._headers: list[str] = []
        self._tables: list[DataTable] = []
        self._charts: list[ChartSpec] = []
        self._footer: str = ""

    def set_title(self, title: str) -> Self:
        self._title = title
        return self

    def add_header(self, text: str) -> Self:
        self._headers.append(text)
        return self

    def add_data_table(self, table: DataTable) -> Self:
        self._tables.append(table)
        return self

    def add_chart(self, chart: ChartSpec) -> Self:
        self._charts.append(chart)
        return self

    def add_footer(self, text: str) -> Self:
        self._footer = text
        return self

    def build(self) -> Report:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        if self._title:
            story.append(Paragraph(self._title, styles["Title"]))
            story.append(Spacer(1, 12))

        for header in self._headers:
            story.append(Paragraph(header, styles["Heading2"]))

        for table in self._tables:
            if table.title:
                story.append(Paragraph(table.title, styles["Heading3"]))

            data = [table.headers] + [[str(v) for v in row] for row in table.rows]
            pdf_table = Table(data)
            pdf_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ]
                )
            )
            story.append(pdf_table)
            story.append(Spacer(1, 12))

        if self._footer:
            story.append(Paragraph(self._footer, styles["Normal"]))

        doc.build(story)
        buf.seek(0)

        return Report(
            title=self._title,
            headers=self._headers,
            data_tables=self._tables,
            charts=self._charts,
            footer=self._footer,
            format=ReportFormat.PDF,
            output=buf.read(),
            content_type="application/pdf",
            filename=f"{self._title or 'report'}.pdf",
        )
