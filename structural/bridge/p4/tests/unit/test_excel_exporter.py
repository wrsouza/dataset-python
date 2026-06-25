"""Unit tests for ExcelFormatExporter."""

from __future__ import annotations

import io

from openpyxl import load_workbook

from exporter.domain.entities import Report
from exporter.infrastructure.excel_exporter import ExcelFormatExporter


def test_serialize_produces_readable_workbook(sample_report: Report) -> None:
    exporter = ExcelFormatExporter()

    payload = exporter.serialize(sample_report)
    workbook = load_workbook(io.BytesIO(payload))
    worksheet = workbook.active

    header = [cell.value for cell in next(worksheet.iter_rows(max_row=1))]
    assert header == ["product", "units"]
    first_data_row = [cell.value for cell in worksheet[2]]
    assert first_data_row == ["Widget", 10]


def test_format_metadata() -> None:
    exporter = ExcelFormatExporter()

    assert exporter.file_extension() == "xlsx"
    assert exporter.format_name() == "Excel"
